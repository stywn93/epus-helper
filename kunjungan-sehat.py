import sys
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import Playwright, sync_playwright


def load_config() -> dict:
    return {
        "BASE_URL": "https://situbondo.epuskesmas.id",
        "EMAIL": "loket-pkm@test.go.id",
        "PASSWORD": "1Sampai9_#",
        "POLI": "KONSELING",
        "DOKTER_DEFAULT": "dr. ANDINI KARTIKA SARI",
        "FILE_EXCEL": "data-pasien.xlsx",
    }


def load_patients_from_excel(filepath: str) -> list[dict]:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    if "nama" not in headers:
        raise ValueError("Required column 'nama' not found in Excel file")

    patients = []
    for row in rows[1:]:
        if not any(row):
            continue
        patient = {}
        for i, header in enumerate(headers):
            if header:
                patient[header] = row[i] if i < len(row) else None
        patients.append(patient)

    wb.close()
    return patients


def login(page, config: dict) -> None:
    page.goto(config["BASE_URL"])
    page.get_by_role("link", name="ePuskesmas").click()
    page.get_by_role("textbox", name="E-mail / No. HP / ID").fill(config["EMAIL"])
    page.get_by_role("textbox", name="E-mail / No. HP / ID").press("Tab")
    page.get_by_role("textbox", name="kata kunci").fill(config["PASSWORD"])
    page.get_by_role("button", name="Login").click()
    page.get_by_role("button", name=" PUSKESMAS SITUBONDO").click()


def register_visit(page, patient_name: str, config: dict) -> None:
    page.get_by_role("button", name="Pendaftaran").click()
    page.get_by_role("link", name="Pasien & KK").click()
    page.get_by_role("textbox", name="Cari Nama").click()
    page.get_by_role("textbox", name="Cari Nama").fill(patient_name)
    page.get_by_role("button", name="Cari").click()
    page.get_by_role("cell").nth(3).dblclick()
    page.get_by_role("cell", name="1", exact=True).click()
    page.get_by_role("cell", name="1", exact=True).click()
    page.get_by_role("link", name="Pendaftaran", exact=True).click()
    page.get_by_role("radio", name="KUNJUNGAN SEHAT").check()
    page.get_by_role("checkbox", name="Kondisi stabil").check()
    page.get_by_role("button", name=" Rawat Jalan").click()
    page.get_by_role("button", name="f  KONSELING").click()
    page.get_by_role("button", name="   dr. ANDINI KARTIKA SARI 0/").click()
    page.get_by_role("button", name="f   dr. RARAS SILVIA GAMA 0/").click()
    page.locator("#button_save").click()


def open_medical_record(page, patient_name: str) -> None:
    page.get_by_role("button", name="Pelayanan").click()
    page.get_by_role("link", name="Medis").click()
    page.locator("i").first.click()
    page.get_by_role("cell", name="23", exact=True).click()
    page.get_by_role("textbox", name="Pencarian").click()
    page.get_by_role("textbox", name="Pencarian").fill(patient_name)
    page.get_by_role("button", name="Cari").click()
    page.get_by_role("cell", name="1", exact=True).dblclick()
    page.get_by_role("button", name="Ok", exact=True).click()
    page.get_by_role("link", name="Anamnesa", exact=True).click()


def main() -> None:
    config = load_config()

    excel_path = config["FILE_EXCEL"]
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]

    try:
        patients = load_patients_from_excel(excel_path)
    except (FileNotFoundError, ValueError) as e:
        print(f"Error loading patients: {e}")
        sys.exit(1)

    print(f"Loaded {len(patients)} patient(s) from {excel_path}")
    for i, p in enumerate(patients, 1):
        print(f"  {i}. {p.get('nama', '(no name)')}")

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        try:
            login(page, config)
            print("Login successful")

            for patient in patients:
                name = patient.get("nama")
                if not name:
                    print("Skipping patient with no name")
                    continue

                print(f"Processing: {name}")
                try:
                    register_visit(page, name, config)
                    print(f"  Visit registered for {name}")
                except Exception as e:
                    print(f"  Failed to register visit for {name}: {e}")

        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
